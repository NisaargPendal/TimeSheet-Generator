from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from datetime import datetime, timedelta
import calendar

# Configuration JSON
STAFF_CONFIG = {
    "nisarg": ["Company1", "Company2"],
    "Deval": ["Company3"],
    "Krutarth": ["Company4"]
}

# Color configurations with exact hex codes
LIGHT_ORANGE = PatternFill(start_color='FCE5CD', end_color='FCE5CD', fill_type='solid')  # #fce5cd
DARK_GRAY = PatternFill(start_color='999999', end_color='999999', fill_type='solid')     # #999999
HEADER_GRAY = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')   # #d9d9d9
HOLIDAY_YELLOW = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid') # #ffff00

# Public Holidays for 2025
HOLIDAYS = {
    datetime(2025, 1, 14): "Makara Sankranti",
    datetime(2025, 1, 26): "Republic Day",
    datetime(2025, 2, 14): "Maha Shivaratri",
    datetime(2025, 3, 3): "Holi",
    datetime(2025, 8, 15): "Independence Day",
    datetime(2025, 8, 28): "Raksha Bandhan",
    datetime(2025, 9, 5): "Janmashtami",
    datetime(2025, 9, 15): "Ganesh Chaturthi",
    datetime(2025, 10, 2): "Gandhi Jayanti",
    datetime(2025, 10, 11): "Dussehra (Vijaya Dashami)",
    datetime(2025, 10, 31): "Diwali",
    datetime(2025, 11, 1): "Vikram Samvat New Year",
    datetime(2025, 11, 2): "Bhai Dooj",
    datetime(2025, 11, 24): "Guru Nanak Jayanti",
    datetime(2025, 12, 25): "Christmas Day"
}

# Border configuration
BORDER_STYLE = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Font configurations - size 10 and Arial
DEFAULT_FONT = Font(name='Arial', size=10)
HEADER_FONT = Font(name='Arial', size=10, bold=True)  # Bold for headers

def get_month_dates(month_num, year=2025):
    """Generate all dates for given month number with holiday information."""
    num_days = calendar.monthrange(year, month_num)[1]

    dates = []
    for day in range(1, num_days + 1):
        date = datetime(year, month_num, day)
        holiday_name = HOLIDAYS.get(date, None)
        dates.append({
            'date': date,
            'day_name': date.strftime('%A'),
            'is_weekend': date.weekday() >= 5,
            'holiday_name': holiday_name
        })
    return dates

def apply_cell_style(cell, fill_color, is_header=False):
    """Apply consistent styling to a cell."""
    cell.font = HEADER_FONT if is_header else DEFAULT_FONT
    cell.fill = fill_color
    cell.border = BORDER_STYLE
    cell.alignment = Alignment(horizontal='center', vertical='center')

def create_excel_sheet(person_name, companies, month_num):
    """Create Excel sheet for a specific person."""
    wb = Workbook()
    ws = wb.active

    # Set up headers
    headers = ['Date', 'Day', 'Holiday', 'Working Hours'] + companies
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        apply_cell_style(cell, HEADER_GRAY, is_header=True)

    # Get dates for the month
    dates = get_month_dates(month_num)

    # Fill data
    for row, date_info in enumerate(dates, 2):
        # Determine fill color based on holiday and weekend status
        if date_info['holiday_name']:
            fill_color = HOLIDAY_YELLOW  # Holiday takes priority
        elif date_info['is_weekend']:
            fill_color = DARK_GRAY
        else:
            fill_color = LIGHT_ORANGE

        # Date column
        date_cell = ws.cell(row=row, column=1)
        date_cell.value = date_info['date'].strftime('%d/%m/%Y')
        apply_cell_style(date_cell, fill_color)

        # Day column
        day_cell = ws.cell(row=row, column=2)
        day_cell.value = date_info['day_name']
        apply_cell_style(day_cell, fill_color)

        # Holiday column
        holiday_cell = ws.cell(row=row, column=3)
        holiday_cell.value = date_info['holiday_name'] if date_info['holiday_name'] else ""
        apply_cell_style(holiday_cell, fill_color)

        # Working Hours column (blank)
        hours_cell = ws.cell(row=row, column=4)
        hours_cell.value = ""
        apply_cell_style(hours_cell, fill_color)

        # Company columns
        for col, _ in enumerate(companies, 5):
            cell = ws.cell(row=row, column=col)
            cell.value = ""
            apply_cell_style(cell, fill_color)

    # Adjust column widths
    ws.column_dimensions['A'].width = 12  # Date column
    ws.column_dimensions['B'].width = 12  # Day column
    ws.column_dimensions['C'].width = 30  # Holiday column
    ws.column_dimensions['D'].width = 15  # Working Hours column

    # Set width for company columns
    for i, _ in enumerate(companies, 5):
        ws.column_dimensions[chr(64 + i)].width = 12

    # Save file with month name
    month_name = calendar.month_name[month_num].lower()
    filename = f"{person_name}-{month_name}.xlsx"
    wb.save(filename)
    return filename

def generate_sheets(month_num):
    """Generate Excel sheets for all staff members."""
    generated_files = []

    for person, companies in STAFF_CONFIG.items():
        filename = create_excel_sheet(person, companies, month_num)
        generated_files.append(filename)

    return generated_files

def main():
    try:
        # Get month number input from user
        month_num = int(input("Enter month number (1-12): ").strip())

        # Validate month number
        if month_num < 1 or month_num > 12:
            raise ValueError("Month number must be between 1 and 12!")

        # Generate sheets
        generated_files = generate_sheets(month_num)

        print("\nGenerated Excel files:")
        for file in generated_files:
            print(f"- {file}")

    except ValueError as e:
        print(f"Error: {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

if __name__ == "__main__":
